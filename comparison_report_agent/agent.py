# comparison_report_agent/agent.py
"""
Comparison Report Sub-Agent
Builds side-by-side competitor comparison tables and Excel workbooks.

Fixes applied:
  - EXCEL_FILE resolved lazily at call time so OUTPUT_DIR env var is always respected
  - Summary sheet now aggregates by unique company (cross-company comparison)
  - Chart reflects per-company totals, not per-run rows
  - Header row styled consistently across all sheets
"""

import os
from datetime import datetime
from google.adk.agents import LlmAgent
from google.adk.models.lite_llm import LiteLlm
from google.adk.tools import FunctionTool
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


def _excel_path() -> str:
    """Resolve Excel path at call time so env var changes are always picked up."""
    out = os.environ.get("MARKET_SCOUT_OUTPUT_DIR", os.path.join(os.getcwd(), "outputs"))
    os.makedirs(out, exist_ok=True)
    return os.path.join(out, "market_scout_data.xlsx")


def _hfill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _border() -> Border:
    thin = Side(style="thin", color="D0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


STATUS_COLORS = {
    "WEEK"      : "C6EFCE",
    "MONTH"     : "FFEB9C",
    "YEAR"      : "DDEBF7",
    "OTHER SOURCES": "F2F2F2",
    "STALE"     : "FFC7CE",
}


def update_excel(all_runs: list) -> str:
    """
    Writes all historical runs to a persistent Excel workbook.

    Sheet 1 — All Features  : every feature row colour-coded by status
    Sheet 2 — By Company    : one row per unique company, aggregated totals + bar chart
    Sheet 3 — Run History   : one row per run (chronological log)

    Returns the Excel file path.
    """
    path = _excel_path()
    wb   = openpyxl.Workbook()

    # ── Sheet 1: All Features ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "All Features"

    hdr1 = ["Run Date", "Company", "Feature", "Category", "Published Date", "Status", "Source URL"]
    for col, h in enumerate(hdr1, 1):
        c       = ws1.cell(row=1, column=col, value=h)
        c.fill  = _hfill("2D1B69")
        c.font  = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
    ws1.row_dimensions[1].height = 22

    row = 2
    for run in all_runs:
        for f in run.get("features", []):
            status = f.get("status", "")
            fill   = _hfill(STATUS_COLORS.get(status, "FFFFFF"))
            vals   = [
                run.get("run_date", ""),
                run.get("company", ""),
                f.get("feature", ""),
                f.get("category", ""),
                f.get("date", "unknown"),
                status,
                f.get("url", ""),
            ]
            for col, val in enumerate(vals, 1):
                c           = ws1.cell(row=row, column=col, value=val)
                c.fill      = fill
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.border    = _border()
                c.font      = Font(size=9)
            row += 1

    for col, width in enumerate([16, 18, 50, 14, 14, 13, 55], 1):
        ws1.column_dimensions[get_column_letter(col)].width = width
    ws1.freeze_panes = "A2"

    # ── Sheet 2: By Company (cross-company comparison + chart) ───────────────
    ws2 = wb.create_sheet("By Company")

    hdr2 = ["Company", "Total Features", "Last 7 Days (WEEK)",
            "Last 30 Days (MONTH)", "Last 365 Days (YEAR)", "Other Sources", "Stale"]
    for col, h in enumerate(hdr2, 1):
        c       = ws2.cell(row=1, column=col, value=h)
        c.fill  = _hfill("4B0082")
        c.font  = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
    ws2.row_dimensions[1].height = 30

    # Aggregate by company across ALL runs
    company_stats: dict[str, dict] = {}
    for run in all_runs:
        comp = run.get("company", "Unknown")
        if comp not in company_stats:
            company_stats[comp] = {"total": 0, "week": 0, "month": 0, "year": 0, "unver": 0, "stale": 0}
        for f in run.get("features", []):
            s = f.get("status", "")
            company_stats[comp]["total"] += 1
            if s == "WEEK":
                company_stats[comp]["week"]  += 1
                company_stats[comp]["month"] += 1
                company_stats[comp]["year"]  += 1
            elif s == "MONTH":
                company_stats[comp]["month"] += 1
                company_stats[comp]["year"]  += 1
            elif s == "YEAR":
                company_stats[comp]["year"]  += 1
            elif s == "OTHER SOURCES":
                company_stats[comp]["unver"] += 1
            elif s == "STALE":
                company_stats[comp]["stale"] += 1

    alt_fills = [_hfill("F3EEFF"), _hfill("EDE7FF")]
    for i, (comp, st) in enumerate(company_stats.items(), 2):
        fill = alt_fills[(i - 2) % 2]
        vals = [comp, st["total"], st["week"], st["month"], st["year"], st["unver"], st["stale"]]
        for col, val in enumerate(vals, 1):
            c           = ws2.cell(row=i, column=col, value=val)
            c.fill      = fill
            c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
            c.border    = _border()
            c.font      = Font(size=10, bold=(col == 1))

    for col, width in enumerate([22, 16, 18, 20, 20, 12, 10], 1):
        ws2.column_dimensions[get_column_letter(col)].width = width
    ws2.freeze_panes = "A2"

    # Bar chart — cross-company comparison
    n_companies = len(company_stats)
    if n_companies > 0:
        chart              = BarChart()
        chart.type         = "col"
        chart.grouping     = "clustered"
        chart.title        = "Feature Coverage by Company"
        chart.y_axis.title = "Feature Count"
        chart.x_axis.title = "Company"
        chart.style        = 10
        chart.width        = 20
        chart.height       = 12

        data = Reference(ws2, min_col=2, max_col=5, min_row=1, max_row=n_companies + 1)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=n_companies + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws2.add_chart(chart, "I2")

    # ── Sheet 3: Run History (chronological log) ─────────────────────────────
    ws3 = wb.create_sheet("Run History")

    hdr3 = ["Run Date", "Company", "Total", "Week", "Month", "Year", "Other Sources"]
    for col, h in enumerate(hdr3, 1):
        c       = ws3.cell(row=1, column=col, value=h)
        c.fill  = _hfill("1F4E79")
        c.font  = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
    ws3.row_dimensions[1].height = 22

    for i, run in enumerate(all_runs, 2):
        feats   = run.get("features", [])
        summary = run.get("summary", {})
        vals    = [
            run.get("run_date", ""),
            run.get("company", ""),
            summary.get("total", len(feats)),
            summary.get("week", 0),
            summary.get("month", 0),
            summary.get("year", 0),
            summary.get("unver", 0),
        ]
        fill = _hfill("EBF5FB") if i % 2 == 0 else _hfill("FFFFFF")
        for col, val in enumerate(vals, 1):
            c           = ws3.cell(row=i, column=col, value=val)
            c.fill      = fill
            c.alignment = Alignment(horizontal="center" if col > 2 else "left", vertical="center")
            c.border    = _border()
            c.font      = Font(size=9)

    for col, width in enumerate([18, 22, 10, 10, 10, 10, 12], 1):
        ws3.column_dimensions[get_column_letter(col)].width = width
    ws3.freeze_panes = "A2"

    wb.save(path)
    return path


def build_comparison_table(runs: list) -> str:
    """
    Builds a markdown side-by-side comparison table for multiple companies.
    Each run dict must contain: company, summary (total/week/month/year/unver).
    """
    if not runs:
        return "_No runs provided for comparison._"

    header = "| Metric |"
    sep    = "|:-------|"
    row_total = "| **Total Features** |"
    row_week  = "| Last 7 Days 🟢 |"
    row_month = "| Last 30 Days 🟡 |"
    row_year  = "| Last 365 Days 🔵 |"
    row_unver = "| Other Sources ⚪ |"

    for run in runs:
        company = run.get("company", "?")
        summary = run.get("summary", {})
        header    += f" **{company}** |"
        sep       += ":----------:|"
        row_total += f" {summary.get('total', 0)} |"
        row_week  += f" {summary.get('week', 0)} |"
        row_month += f" {summary.get('month', 0)} |"
        row_year  += f" {summary.get('year', 0)} |"
        row_unver += f" {summary.get('unver', 0)} |"

    return "\n".join([header, sep, row_total, row_week, row_month, row_year, row_unver])


excel_tool      = FunctionTool(func=update_excel)
comparison_tool = FunctionTool(func=build_comparison_table)

comparison_report_agent = LlmAgent(
    name="comparison_report_agent",
    model=LiteLlm(model="groq/llama-3.1-8b-instant"),
    description="Builds Excel workbooks and comparison tables for multi-company runs.",
    instruction=(
        "You are a Comparison Report Agent. "
        "When given a list of run dicts: "
        "1. Call update_excel(all_runs) to persist data to Excel. "
        "2. Call build_comparison_table(runs) to produce a markdown comparison table. "
        "Return both outputs exactly as received."
    ),
    tools=[excel_tool, comparison_tool],
)
